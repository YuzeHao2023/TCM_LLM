import openpyxl
import json
import random


alias_keywords=['{}有其他别名嘛','{}别名是什么','{}还有其他称呼','{}别称','{}别名','{}还可以怎么叫','{}其他名字','{}哪些别称']
smell_keywords=['{}什么气味','{}什么味道','{}闻起来怎么样','{}什么味的','{}吃起来苦嘛','{}吃了是什么味','{}有毒嘛','{}有毒性嘛','{}的药性是什么']
cure_keywords=['{}可以治什么','{}治哪些病','{}可以治什么症状','{}有什么好处', '{}有什么益处', '{}有何益处','{}用来做啥', '{}用来作甚','{}治愈啥', '{}主治啥', '{}主治什么', '{}有什么用', '{}有何用']
part_keywords=['{}属于什么部类','{}属于什么部','{}什么部类', '{}什么部', '{}哪个部类', '{}哪个部']
symptom_keywords=['{}怎么治','{}有什么药方','{}有什么方','{}中医怎么治','{}怎么处理','{}怎么弄','{}怎么搞','{}有啥法子','{}有什么办法']
repeat_times = 1

def process_excel_to_json(input_files, symptom_files, output_file, test_file):
    # Initialize the output data structure
    system_value = "您是一位非常专业的的中医药学教授。您始终根据提问者的问题提供准确、全面和详细的答案。"
    output_data = []
    test_data = []
    for r_t in range(repeat_times):
        hellodata = [
            {
                "conversation": [
                    {
                        "system": system_value,
                        "input": "请做一下自我介绍",
                        "output": "您好，我是中医药知识问答小助手，您可以向我提问一些中药的信息或者对某些症状的药方，我会尽我所知回答您"
                    }
                ]
            },
            {
                "conversation": [
                    {
                        "system": system_value,
                        "input": "请介绍一下你自己",
                        "output": "您好，我是中医药知识问答小助手，您可以向我提问一些中药的信息或者对某些症状的药方，我会尽我所知回答您"
                    }
                ]
            },
            {
                "conversation": [
                    {
                        "system": system_value,
                        "input": "你好",
                        "output": "您好，我是中医药知识问答小助手，您可以向我提问一些中药的信息或者对某些症状的药方，我会尽我所知回答您"
                    }
                ]
            },
            {
                "conversation": [
                    {
                        "system": system_value,
                        "input": "您好",
                        "output": "您好，我是中医药知识问答小助手，您可以向我提问一些中药的信息或者对某些症状的药方，我会尽我所知回答您"
                    }
                ]
            },
            {
                "conversation": [
                    {
                        "system": system_value,
                        "input": "你好，请介绍一下自己",
                        "output": "您好，我是中医药知识问答小助手，您可以向我提问一些中药的信息或者对某些症状的药方，我会尽我所知回答您"
                    },
                ]
            },
            {
                "conversation": [
                    {
                        "system": system_value,
                        "input": "你是谁",
                        "output": "您好，我是中医药知识问答小助手，您可以向我提问一些中药的信息或者对某些症状的药方，我会尽我所知回答您"
                    },
                ]
            },
            {
                "conversation": [
                    {
                        "system": system_value,
                        "input": "你是",
                        "output": "您好，我是中医药知识问答小助手，您可以向我提问一些中药的信息或者对某些症状的药方，我会尽我所知回答您"
                    },
                ]
            },
            {
                "conversation": [
                    {
                        "system": system_value,
                        "input": "你哪位",
                        "output": "您好，我是中医药知识问答小助手，您可以向我提问一些中药的信息或者对某些症状的药方，我会尽我所知回答您"
                    },
                ]
            }
        ]
        output_data += hellodata
        for input_file in input_files:
            # Load the workbook
            wb = openpyxl.load_workbook(input_file)

            # Select the "DrugQA" sheet
            sheet = wb["Sheet"]

            # Iterate through each row in column A and D
            for row in sheet.iter_rows(min_row=2, max_col=5, values_only=True):

                # Create the conversation dictionary
                print(row[0],len(row))
                if len(row[1])>0:
                    random_list = random.sample([k for k in range(len(part_keywords))], random.randint(int(len(part_keywords)*2/3),len(part_keywords)))
                    test_list = list(set([k for k in range(len(part_keywords))])-set(random_list))
                    for i in random_list:
                        conversation = {
                            "system": system_value,
                            "input": part_keywords[i].format(row[0]),
                            #"output": {"name":row[1],"question_type":"part","answer":row[0]}
                            #"output": json.dumps({"name":row[0],"question_type":"part","answer":row[1]},ensure_ascii=False)
                            "output": row[0]+"所属的部是"+row[1]
                        }

                        # Append the conversation to the output data
                        output_data.append({"conversation": [conversation]})

                    for i in test_list:
                        qa = {
                            "question": part_keywords[i].format(row[0]),
                            "answer": row[0]+"所属的部是"+row[1]
                        }

                        # Append the conversation to the output data
                        test_data.append(json.dumps(qa,ensure_ascii=False))

                if len(row)>=3:
                    if row[2]:
                        random_list = random.sample([k for k in range(len(alias_keywords))], random.randint(int(len(alias_keywords)*2/3),len(alias_keywords)))
                        test_list = list(set([k for k in range(len(alias_keywords))])-set(random_list))
                        for i in random_list:
                            conversation = {
                                "system": system_value,
                                "input": alias_keywords[i].format(row[0]),
                                #"output": {"name":row[1],"question_type":"alias","answer":row[2]}
                                #"output": json.dumps({"name":row[0],"question_type":"alias","answer":row[2]},ensure_ascii=False)
                                "output": row[0]+"的名称解释或者别名是"+row[2]
                            }

                            # Append the conversation to the output data
                            output_data.append({"conversation": [conversation]})

                        for i in test_list:
                            qa = {
                                "question": alias_keywords[i].format(row[0]),
                                "answer": row[0]+"的名称解释或者别名是"+row[2]
                            }

                            # Append the conversation to the output data
                            test_data.append(json.dumps(qa,ensure_ascii=False))

                if len(row)>=4:
                    if row[3]:
                        random_list = random.sample([k for k in range(len(smell_keywords))], random.randint(int(len(smell_keywords)*2/3),len(smell_keywords)))
                        test_list = list(set([k for k in range(len(smell_keywords))])-set(random_list))
                        for i in random_list:
                            conversation = {
                                "system": system_value,
                                "input": smell_keywords[i].format(row[0]),
                                #"output": {"name":row[1],"question_type":"smell","answer":row[3]}
                                #"output": json.dumps({"name":row[0],"question_type":"smell","answer":row[3]},ensure_ascii=False)
                                "output": row[0]+"的气味是"+row[3]
                            }

                            # Append the conversation to the output data
                            output_data.append({"conversation": [conversation]})
                        for i in test_list:
                            qa = {
                                "question": smell_keywords[i].format(row[0]),
                                "answer": row[0]+"的气味是"+row[3]
                            }

                            # Append the conversation to the output data
                            test_data.append(json.dumps(qa,ensure_ascii=False))

                if len(row)>=5:
                    if row[4]:
                        random_list = random.sample([k for k in range(len(cure_keywords))], random.randint(int(len(cure_keywords)*2/3),len(cure_keywords)))
                        test_list = list(set([k for k in range(len(cure_keywords))])-set(random_list))
                        for i in random_list:
                            conversation = {
                                "system": system_value,
                                "input": cure_keywords[i].format(row[0]),
                                #"output": {"name":row[1],"question_type":"cure","answer":row[4]}
                                #"output": json.dumps({"name":row[0],"question_type":"cure","answer":row[4]},ensure_ascii=False)
                                "output": row[0]+"的功效是"+row[4]
                            }

                            # Append the conversation to the output data
                            output_data.append({"conversation": [conversation]})

                        for i in test_list:
                            qa = {
                                "question": cure_keywords[i].format(row[0]),
                                "answer": row[0]+"的功效是"+row[4]
                            }

                            # Append the conversation to the output data
                            test_data.append(json.dumps(qa,ensure_ascii=False))

        for symptom_file in symptom_files:
            # Load the workbook
            wb = openpyxl.load_workbook(symptom_file)

            # Select the "DrugQA" sheet
            sheet = wb["Sheet"]

            
            symptomdict = {}
            # Iterate through each row in column A and D
            for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
                if row[1] not in symptomdict.keys():
                    symptomdict[row[1]] = [row[2]]
                else:
                    symptomdict[row[1]].append(row[2])

            for key in symptomdict.keys():
                system_value = "您是一位非常专业的的中医药学教授。您始终根据提问者的问题提供准确、全面和详细的答案。"

                # Create the conversation dictionary
                symptom_list = symptomdict[key]
                symptom_method = '\n'.join(['药方{}:'.format(j+1)+symptom_list[j] for j in range(len(symptom_list))])
                if len(symptom_method)>0:
                    for i in random.sample([k for k in range(len(symptom_keywords))], random.randint(int(len(symptom_keywords)*2/3),len(symptom_keywords))):
                        conversation = {
                            "system": system_value,
                            "input": symptom_keywords[i].format(key),
                            #"output": {"name":row[1],"question_type":"part","answer":row[0]}
                            #"output": json.dumps({"name":key,"question_type":"method","answer":symptom_method},ensure_ascii=False)
                            "output": '对'+key+"有以下药方："+str(symptom_method.split("\n"))
                        }

                        # Append the conversation to the output data
                        output_data.append({"conversation": [conversation]})

    # Write the output data to a JSON file
    # Shuffle the data randomly
    random.shuffle(output_data)
    random.shuffle(output_data)
    random.shuffle(output_data)

    random.shuffle(test_data)
    random.shuffle(test_data)
    random.shuffle(test_data)

    
    with open(output_file, 'w', encoding='utf-8') as json_file:
        json.dump(output_data, json_file, indent=4,ensure_ascii=False)

    with open(test_file, 'w', encoding='utf-8') as json_file:
        json_file.write('\n'.join(test_data))

    print(f"Conversion complete. Output written to {output_file}")

# Replace 'MedQA2019.xlsx' and 'output.jsonl' with your actual input and output file names
process_excel_to_json(['./data/xlsx_new/ChineseMedicalNew.xlsx'],['./data/xlsx_new/SymptomNew.xlsx'], './data/jsonl3/train.jsonl','./data/jsonl3/test.jsonl')
